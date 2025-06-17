#!/usr/bin/env python3
# coding=utf-8

import csv
import os
import re
import urllib.request

import openpyxl

from DataGenerator import DataGenerator


class NyseDataGenerator(DataGenerator):

    # private static variables
    __WHITELIST = ["NYSE", "ARCA", "BATSZ", "AMEX"]
    __SRC_DATA_XLSX = 'symbolsNYSE.xlsx'
    __MY_CSV = "symbolsNYSE.csv"
    __URL = 'https://www.nyse.com/publicdocs/nyse/symbols/Symbol_Distribution.xlsx'

    @staticmethod
    def _download_file(url: str, destination: str) -> None:
        """
        :param url: url where the file is downloaded from
        :param destination: the name of the file in which the downloaded file is saved
        """
        urllib.request.urlretrieve(url, destination)

    @staticmethod
    def _xlsx_to_csv(source: str, destination: str) -> None:
        """
        :param source: path to the source XLSX file
        :param destination: path to the resulting CSV file
        """
        workbook = openpyxl.load_workbook(source, read_only=True)

        sheet_name = "ARCX"
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' was not found in '{source}'")
            sheet = workbook[sheet_name]

            with open(destination, 'w', newline='', encoding='utf-8') as csv_file:
                csv_writer = csv.writer(csv_file)
                for row in sheet.iter_rows(values_only=True):
                    csv_writer.writerow(row)
        finally:
            workbook.close()

    @staticmethod
    def __get_mic_code(market: str) -> str:
        """
        :param market: market from the exchange
        :return: mic-code for symbolinfo
        """
        if market == "NYSE":
            return "XNYS"
        if market == "ARCA":
            return "ARCX"
        if market == "AMEX":
            return "XASE"
        if market == "BATSZ":
            return "BATS"
        raise AttributeError(f"Unknown market {market}")

    @staticmethod
    def __fix_description(description: str) -> str:
        """
        :param description: description of the symbol from the exchange
        :return: description for symbolinfo
        """
        if "�" in description:
            return description.replace("�", " ")

        if ";" in description:
            return f"{description.split(';')[0]} -{description.split(';')[1]}"

        return description

    @staticmethod
    def __get_tv_symbol(symbol: str) -> str:
        """
        :param symbol: symbol from the exchange
        :return: tv-symbol for symbolinfo
        """
        tv_symbol = re.match(
            r'(?P<pr1>.*)\s(?P<pr2>[PR])R(?P<pr3>.)*|'
            r'(?P<dotsym1>.*)\s(?P<dotsym2>.$)|'
            r'(?P<ws1>.*)\sWS(?P<ws2>.)*|'
            r'(?P<rt1>.*)\sRT', symbol
        )
        if tv_symbol and tv_symbol.group('pr1') is not None:
            return (tv_symbol.group('pr1') + '/' + tv_symbol.group('pr2')) + (
                tv_symbol.group('pr3') if tv_symbol.group('pr3') is not None else '')

        if tv_symbol and tv_symbol.group('dotsym1') is not None:
            return tv_symbol.group('dotsym1') + '.' + tv_symbol.group('dotsym2')

        if tv_symbol and tv_symbol.group('ws1') is not None:
            return (tv_symbol.group('ws1') + '/W') + (tv_symbol.group('ws2') if tv_symbol.group('ws2') is not None else '')

        if tv_symbol and tv_symbol.group('rt1') is not None:
            return tv_symbol.group('rt1') + '/R'

        return symbol

    def _get_symbols_info(self, my_csv: str) -> tuple:
        """
        :param my_csv: csv file name
        :return: 2 lists with symbols, descriptions and mic-codes for symbolinfo
        """
        nyse_symbol_list = []
        amex_symbol_list = []

        with open(my_csv, encoding='utf-8') as symbols:
            file = csv.reader(symbols, delimiter=",")
            for i in file:
                if i[5] in NyseDataGenerator.__WHITELIST:
                    description = self.__fix_description(i[0])
                    tv_sym = self.__get_tv_symbol(i[1])
                    mic = self.__get_mic_code(i[5])

                    if i[5] == "NYSE":
                        nyse_symbol_list.append(f"{tv_sym};{description};{mic}")
                    else:
                        amex_symbol_list.append(f"{tv_sym};{description};{mic}")

        return nyse_symbol_list, amex_symbol_list

    def generate(self) -> list[str]:
        try:
            self._download_file(NyseDataGenerator.__URL, NyseDataGenerator.__SRC_DATA_XLSX)
            self._xlsx_to_csv(NyseDataGenerator.__SRC_DATA_XLSX, NyseDataGenerator.__MY_CSV)
            os.remove(NyseDataGenerator.__SRC_DATA_XLSX)

            nyse_symbol_list, amex_symbol_list = self._get_symbols_info(NyseDataGenerator.__MY_CSV)
            os.remove(NyseDataGenerator.__MY_CSV)

            with open("nyse_data.csv", "w", encoding="utf-8") as file:
                file.write("tv-symbol;description;mic-code\n")
                for i in nyse_symbol_list:
                    file.write(f"{i}\n")

            with open("amex_data.csv", "w", encoding="utf-8") as file:
                file.write("tv-symbol;description;mic-code\n")
                for i in amex_symbol_list:
                    file.write(f"{i}\n")
        except (OSError, AttributeError, KeyError, ValueError, NotImplementedError) as e:
            self._logger.error(e)
            raise e
        return ["nyse_data.csv", "amex_data.csv"]


if __name__ == "__main__":
    try:
        NyseDataGenerator().generate()
        exit(0)
    except (OSError, AttributeError, KeyError, ValueError, NotImplementedError):
        exit(1)
