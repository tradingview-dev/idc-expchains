#!/usr/bin/env python3
# coding=utf-8

import os
import urllib.request

import openpyxl

from DataGenerator import DataGenerator


class NASDAQGIDSDataGenerator(DataGenerator):

    # private static variables
    __DESTINATION = 'nasdaq_gids_symbols.xlsx'
    __MY_CSV = "nasdaq_gids_symbols.csv"
    __URLs = ['https://indexes.nasdaqomx.com/Index/ExportDirectory', 'https://indexes.nasdaqomx.com/Home/ExportGidsDirectory']

    @staticmethod
    def download_file(url: str, destination: str) -> None:
        """
        Downloads a file from the specified URL and saves it to the destination.

        :param url: URL where the file is downloaded from
        :param destination: The path where the downloaded file will be saved
        """
        urllib.request.urlretrieve(url, destination)

    @staticmethod
    def get_indices_from_xlsx(filename: str) -> set:
        """
        Extracts and returns a sorted list of symbols from a xlsx file.

        :param filename: Path to the xlsx file
        :return: Sorted list of indices
        """
        gids_indices = set()
        workbook = openpyxl.load_workbook(filename, read_only=True)

        sheet_name = "Index Directory"
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' was not found in '{filename}'")
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows(values_only=True):
                if len(row) == 3:
                    if row[2] != "I":
                        continue
                else:
                    if (row[7] == "" and row[6] == "SandP") or row[0] == "Symbol":
                        continue
                gids_indices.add(row[0])
        finally:
            workbook.close()

        return gids_indices

    @staticmethod
    def write_to_csv(dst: str, indices: list) -> None:
        """
        Writes a list of indices to a CSV file.

        :param dst: Destination
        :param indices: Sorted list of indices to save in the file
        """
        with open(dst, 'w', newline='', encoding='utf-8') as csv_file:
            csv_file.write("tv-symbol\n")
            for index in indices:
                csv_file.write(index + "\n")

    def generate(self) -> list[str]:
        gids_indices = set()
        for url in self.__URLs:
            try:
                self.download_file(url, self.__DESTINATION)
                gids_indices.update(self.get_indices_from_xlsx(self.__DESTINATION))
            except (OSError, ValueError, IndexError) as e:
                self._logger.error(e)
                raise e
        try:
            self.write_to_csv(self.__MY_CSV, sorted(gids_indices))
            os.remove(self.__DESTINATION)
        except OSError as e:
            self._logger.error(e)
            raise e

        return [self.__MY_CSV]


if __name__ == "__main__":
    try:
        NASDAQGIDSDataGenerator().generate()
        exit(0)
    except (OSError, ValueError, IndexError):
        exit(1)
